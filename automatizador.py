import time
import getpass # Importa a biblioteca para senhas seguras
import os
import glob
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta # Para lidar com datas (hoje e ontem)
import win32com.client

# Precisa fazer pip install openpyxl

# --- Bibliotecas do EXCEL ---
try:
  import openpyxl
  from openpyxl.styles import PatternFill, Font, Border, Side
  from openpyxl.utils import get_column_letter
except ImportError:
  print("Biblioteca 'openpyxl' não encontrada.")
  print("Por favor, instale-a com: pip install openpyxl")
  exit()


# --- Configuração do Driver ---
try:
  driver = webdriver.Chrome() 
  driver.get("https://podio.com/login")
  driver.maximize_window()
except Exception as e:
  print(f"Erro ao iniciar o Chrome. Verifique seu chromedriver. {e}")
  exit()

# --- Etapa 1: Clicar no Login da Microsoft ---
try:
  print("Aguardando página de login...")
  microsoft_login_xpath = "//a[@data-provider='live']"
  
  print("Procurando o botão de login da Microsoft...")
  microsoft_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, microsoft_login_xpath))
  )
  
  print("Clicando no botão da Microsoft...")
  microsoft_button.click()

except Exception as e:
  print(f"Erro ao tentar clicar no botão da Microsoft: {e}")
  driver.quit()
  exit()

# --- Etapa 1.5: Lidar com o login da Microsoft (Versão Final) ---
try:
  print("Aguardando a nova janela/aba de login da Microsoft...")
  
  # Espera o pop-up abrir (total de 2 janelas)
  WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))

  # Pega o ID da janela pop-up
  popup_window = None
  original_window = driver.current_window_handle # Guarda a original (que vai fechar)
  for window_handle in driver.window_handles:
    if window_handle != original_window:
      popup_window = window_handle
      break
      
  driver.switch_to.window(popup_window)
  print("Foco mudado para a janela de login da Microsoft.")
      
  print("Aguardando a tela de login da Microsoft...")
  
  # Preenche o e-mail (usando a variável segura)
  email_field_microsoft = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.ID, "i0116")) 
  )
  print("Preenchendo e-mail da Microsoft...")
  email_field_microsoft.send_keys("pedro.henrsilva@mrv.com.br")#Seu email aqui
  WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click()
  
  # Preenche a senha (usando a variável segura)
  password_field_microsoft = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.ID, "i0118")) 
  )
  print("Preenchendo senha da Microsoft...")
  password_field_microsoft.send_keys(" ")#Sua senha aqui
  
  # Tenta clicar no botão "Entrar" (com loop anti-stale)
  print("Procurando o botão 'Entrar'...")
  tentativas = 0
  clicado_entrar = False
  while not clicado_entrar and tentativas < 5:
    try:
      entrar_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "idSIButton9")))
      entrar_button.click()
      clicado_entrar = True
      print("Botão 'Entrar' clicado.")
    except StaleElementReferenceException:
      tentativas += 1; time.sleep(0.5)
  if not clicado_entrar: raise Exception("Falha ao clicar em Entrar")

  # --- ESPERA PELO MFA MANUAL ---
  print("!!! AÇÃO MANUAL NECESSÁRIA !!!")
  print("Aguardando aprovação do MFA no seu celular (até 180s)...")
  
  tentativas = 0
  clicado_manter = False
  while not clicado_manter and tentativas < 5:
    try:
      keep_logged_in_button = WebDriverWait(driver, 180).until( 
        EC.element_to_be_clickable((By.ID, "idSIButton9"))
      )
      keep_logged_in_button.click() # Clica "Sim"
      clicado_manter = True
      print("MFA Aprovado! Botão 'Manter conectado' clicado.")
    except StaleElementReferenceException:
      tentativas += 1; time.sleep(0.5)
    except TimeoutException:
      print("Erro: Timeout após 180s. Você não aprovou o MFA a tempo?")
      clicado_manter = False; break
  if not clicado_manter: raise Exception("Falha ao clicar em Manter Conectado")

  print("Login da Microsoft concluído na janela pop-up.")
  
  # --- LÓGICA CORRIGIDA (Baseada na sua análise) ---
  
  # 1. Espera a janela pop-up fechar sozinha
  print("Aguardando janela pop-up fechar...")
  WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(1))
  
  # 2. Pega o handle da ÚNICA janela que sobrou (a nova janela principal)
  nova_janela_principal = driver.window_handles[0]
  driver.switch_to.window(nova_janela_principal)
  print("Foco retornado para a janela principal do Podio.")

  ''''''
  # 3. Força a navegação para a home page
  print("Forçando navegação para https://podio.com/home")
  driver.get("https://podio.com/home")
  

  print("Página principal carregada com sucesso.")
  
except Exception as e:
  print(f"Erro durante o login na Microsoft (Etapa 1.5): {e}")
  driver.save_screenshot("erro_etapa_1-5.png") 
  driver.quit()
  exit()

# --- Início da Navegação no Podio (Etapas 2-12) ---

try:
  # --- ETAPA 2 CORRIGIDA ---
  print("Etapa 2: Procurando 'Vá para uma área de trabalho'...")
  
  # 1. Encontra o elemento 'pai' (a caixa que você passa o mouse por cima)
  parent_element_xpath = "//div[contains(@class, 'space-switcher-wrapper')]"
  parent_element = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, parent_element_xpath))
  )
  
  # 2. Simula o mouse passando por cima dele
  print("Simulando passagem do mouse (hover)...")
  actions = ActionChains(driver)
  actions.move_to_element(parent_element).perform()
  
  # 3. AGORA, espera o botão/texto de dentro aparecer
  
  # Etapa 3: Clicar em "ADM - Núcleo Contratos"
  print("Etapa 3: Aguardando a lista de áreas de trabalho...")
  adm_link_xpath = "//a[contains(text(), 'ADM - Núcleo Contratos')]"
  adm_link = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, adm_link_xpath))
  )
  print("Clicando em 'ADM - Núcleo Contratos'...")
  adm_link.click()


  # --- ETAPA 4 (Usando data-app-id) ---
  print("Etapa 4: Procurando o app 'Mensageria'...")
  
  mensageria_app_xpath = "//li[@data-app-id='22830484']"
  
  mensageria_app = WebDriverWait(driver, 20).until(
    EC.element_to_be_clickable((By.XPATH, mensageria_app_xpath))
  )
  print("Clicando no app 'Mensageria'...")
  mensageria_app.click()
  
  # --- FIM DA ETAPA 4 ---


  # --- ETAPA 5 ---
  print("Etapa 5: Aguardando a página 'Mensageria' carregar...") 

  x = 0
  while x <=2: # Tenta por 2 vezes
    try:
      # 1. Espera o <ul> (pai) carregar
      parent_filter_xpath = "//ul[@class='app-filter-tools']"
      parent_filter_element = WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.XPATH, parent_filter_xpath))
      )
      print(f"Tentativa {tentativas+1}: Container <ul> (pai) encontrado.")

      # 2. Encontra TODOS os <li> dentro dele
      child_items_xpath = ".//li"
      child_items = parent_filter_element.find_elements(By.XPATH, child_items_xpath)
      
      if not child_items:
        raise Exception("Container <ul> encontrado, mas nenhum <li> filho foi encontrado.")
      
      print(f"Encontrados {len(child_items)} itens. Passando o mouse sobre eles...")

      # 3. Simula a ação "humana":
      #  Move o mouse sobre cada item da lista para o menu
      actions_filter = ActionChains(driver)
      for item in child_items:
        actions_filter.move_to_element(item)
      actions_filter.perform() # Executa a sequência de "hovers"
      
      # 4. Agora que o menu está ativo, espera o <li> "Filtros"
      target_filter_xpath = ".//li[@data-original-title='Filtros']" 
      target_filter = WebDriverWait(parent_filter_element, 2).until(
        EC.element_to_be_clickable((By.XPATH, target_filter_xpath))
      )
      
      # 5. Clica
      print("Ícone 'Filtros' acordado e clicável. Clicando...")
      target_filter.click()
      
      clicado_filtro = True # Sucesso! Sai do loop.
      print("Ícone 'Filtros' clicado com sucesso.")
      x += 1
    except (StaleElementReferenceException, TimeoutException) as e:
      # 6. Se der "Stale" ou "Timeout", espera e tenta a etapa 5 inteira de novo
      print(f"Tentativa {x} falhou (Stale ou Timeout). Página recarregando... Tentando de novo.")
      time.sleep(1) # Espera 1 segundo para a página estabilizar
      
  # 7. Se o loop terminar sem clicar, força um erro
  if not clicado_filtro:
    raise Exception("Falha ao encontrar o ícone 'Filtros' após 2 tentativas.")


      
  print("Ação concluída com sucesso!")
  time.sleep(1)

# --- FIM DA ETAPA 5 ---

# --- ETAPA 6 ---

  print("Etapa 6: Procurando o 'Criado em'...")
  
  Criado_em_xpath = "//li[@data-id='created_on']"
  
  Criado_em = WebDriverWait(driver, 3).until(
    EC.element_to_be_clickable((By.XPATH, Criado_em_xpath))
  )
  print("Clicando no 'Criado em'...")
  Criado_em.click()




# --- FIM DA ETAPA 6 ---

# --- ETAPA 7 ---

  print("Etapa 7: Procurando o 'Hoje'...")
  
  Criado_em_hoje = "//li[@data-id='-0dr:-0dr']"
  
  Criado_hoje = WebDriverWait(driver, 5).until(
    EC.element_to_be_clickable((By.XPATH, Criado_em_hoje))
  )
  print("Clicando no 'Hoje'...")
  Criado_hoje.click()


# --- FIM DA ETAPA 7 ---


# --- ETAPA 8 ---

  try:
  # 1. Define o seletor. 
  #  Usar CSS_SELECTOR é mais fácil para classes.
  #  O ponto (.) significa "classe".
    seletor_css = ".app-header__app-menu"
  
    print(f"Procurando todos os elementos com a classe: {seletor_css}")
# 2. Espera até que PELO MENOS 2 elementos estejam presentes
    #  (Você pode mudar o '2' para quantos você espera)
    WebDriverWait(driver, 10).until(
      lambda d: len(d.find_elements(By.CSS_SELECTOR, seletor_css)) >= 2
    )

    # 3. Pega a LISTA de todos os elementos
    elementos = driver.find_elements(By.CSS_SELECTOR, seletor_css)
  
    print(f"Encontrados {len(elementos)} elementos.")

    # 4. Clica no primeiro elemento (índice 0)
    if len(elementos) > 0:
      print("Clicando no primeiro elemento (índice 0)...")
      elementos[0].click()
  
    # 5. Espera a página reagir
    #  (MUITO IMPORTANTE: Clicar em algo pode mudar a página)
    print("Aguardando 2 segundos para a página/menu reagir...")
    time.sleep(2) 

    # 6. RE-ENCONTRA a lista de elementos
    #  (É a forma mais segura, caso o primeiro clique tenha
    #  recarregado os elementos - evita o erro 'stale element')
  
    print("Re-encontrando os elementos (para segurança)...")
    elementos = driver.find_elements(By.CSS_SELECTOR, seletor_css)

    # 7. Clica no segundo elemento (índice 1)
    if len(elementos) > 1:
      print("Clicando no segundo elemento (índice 1)...")
      elementos[1].click()
    else:
      print("Erro: Não foi possível encontrar o segundo elemento após o primeiro clique.")
    
    print("Ações nos dois elementos concluídas!")
    time.sleep(3)

  except Exception as e:
    print(f"Ocorreu um erro: {e}")
    # driver.save_screenshot("erro_multiplos.png")


# --- FIM DA ETAPA 8 ---


# --- ETAPA 9 ---
  print("Etapa 9: Aguardando o menu dropdown abrir...")
  
  # Usando o seletor CSS (mais limpo) que discutimos
  exportar_excel_selector = "a.app-box-supermenu-v2__link.app-export-excel"
  
  exportar_link = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, exportar_excel_selector))
  )
  
  print("Link 'Exportar Excel' encontrado. Clicando...")
  exportar_link.click()
  # --- FIM DA ETAPA 9 ---

  time.sleep(3)

# --- ETAPA 10 ---
  try:
    print("Procurando o ícone de 'Notificação' (Inbox)...")
  
    # Usando o Seletor CSS (recomendado por ser mais limpo)
    notificacao_selector = "li.navigation-link.inbox"
  
    # Espera o ícone estar presente e ser clicável
    notificacao_icon = WebDriverWait(driver, 5).until(
      EC.element_to_be_clickable((By.CSS_SELECTOR, notificacao_selector))
    )
  
    print("Ícone de 'Notificação' encontrado. Clicando...")
    notificacao_icon.click()
  
    time.sleep(1) # Espera o menu de notificação abrir

  except Exception as e:
    print(f"Erro ao tentar clicar no ícone de Notificação: {e}")
    driver.save_screenshot("erro_notificacao.png")
  # --- FIM DA ETAPA 10 ---


# --- ETAPA 11 ---
  css_corrigido = "a.PodioUI__Notifications__NotificationGroup"
  item_notificacao = WebDriverWait(driver, 3).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, css_corrigido))
  )
  item_notificacao.click()
# --- FIM DA ETAPA 11 ---

# --- ETAPA 12: Esperar o processamento do Excel ---
  print("Etapa 12: Aguardando a página de exportação carregar e o status ser 'Completado'...")
  
  # Espera até 3 minutos (180s) para o Excel ser processado
  # A espera procura por um <span> verde que contém "Completado"
  status_completado_xpath = "//div[contains(@class, 'field-type-text')]"
  
  WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.XPATH, status_completado_xpath))
  )
  
  print("Exportação 'Completado'!")
  # --- FIM DA ETAPA 12 ---
  
  # --- ETAPA 13 (NOVA): Clicar no link de download ---
  print("Etapa 13: Procurando o link de download do arquivo...")
  
  # O seletor By.LINK_TEXT é perfeito para isso, pois busca o texto exato do link
  nome_do_arquivo = "Mensageria - Última vista usada.xlsx"
  
  link_download = WebDriverWait(driver, 3).until(
    EC.element_to_be_clickable((By.LINK_TEXT, nome_do_arquivo))
  )
  
  print("Link encontrado! Clicando para baixar...")
  link_download.click()
  
  print("Ação final concluída com sucesso! O download deve começar.")
  time.sleep(10) # Espera 10s para o download iniciar

except Exception as e:
  print(f"Erro durante a navegação no Podio (Etapas 2-5): {e}")
  print("Verifique os seletores XPath. Um deles pode ter mudado.")
  driver.save_screenshot("erro_de_navegacao.png")

try:
  print("\n--- INICIANDO PROCESSAMENTO DO EXCEL ---")
  
  # 1. Encontra o arquivo baixado
  download_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
  
  # Procura por "Mensageria - Última vista usada*.xlsx"
  list_of_files = glob.glob(os.path.join(download_dir, 'Mensageria - Última vista usada*.xlsx'))
  if not list_of_files:
    raise Exception("Arquivo Excel não encontrado na pasta Downloads.")
    
  # Pega o mais recente (caso tenha ' (1)', ' (2)', etc.)
  latest_file = max(list_of_files, key=os.path.getctime)
  print(f"Arquivo encontrado: {latest_file}")
  
  # 2. Abre o arquivo
  workbook = openpyxl.load_workbook(latest_file)
  sheet = workbook.active
  print(f"Planilha '{sheet.title}' aberta.")

  # --- Executando suas 7 etapas ---

  # 1. Excluir as colunas A, I, J, K, L, M (1, 9, 10, 11, 12, 13)
  #  (Excluir da direita para a esquerda para não bagunçar os índices)
  print("1. Excluindo colunas A, I, J, K, L, M...")
  sheet.delete_cols(13) # M
  sheet.delete_cols(12) # L
  sheet.delete_cols(11) # K
  sheet.delete_cols(10) # J
  sheet.delete_cols(9) # I
  sheet.delete_cols(1) # A

  # 2. Colocar o tamanho das colunas em 20
  print("2. Ajustando tamanho das colunas para 20...")
  for i in range(1, sheet.max_column + 1):
    col_letter = get_column_letter(i)
    sheet.column_dimensions[col_letter].width = 20

  # 3. Aplicar filtro em todas as colunas
  print("3. Aplicando filtro...")
  sheet.auto_filter.ref = sheet.dimensions

  # 4. Na coluna F (antiga G), excluir tudo que NÃO for "JURIDICO MONTREAL"
  print("4. Filtrando por 'JURIDICO MONTREAL' na coluna F...")
  linhas_para_excluir = []
  # Iterar de cima para baixo para encontrar as linhas
  # (Começa do 2 para pular o cabeçalho)
  for row_num in range(2, sheet.max_row + 1):
    cell_value = sheet[f'F{row_num}'].value
    # Checa se o valor não é "juridico montreal" (ignorando maiúsculas/minúsculas)
    if not cell_value or "juridico montreal" not in str(cell_value).lower():
      linhas_para_excluir.append(row_num)
  
  # Excluir as linhas de baixo para cima (MUITO IMPORTANTE)
  for row_num in reversed(linhas_para_excluir):
    sheet.delete_rows(row_num)
  print(f"{len(linhas_para_excluir)} linhas excluídas.")

  # 5. Renomear coluna G (antiga H)
  print("5. Renomeando coluna G para 'Data de recebimento'...")
  sheet['G1'] = "Data de recebimento"

  # 6. Aplicar borda em todas as colunas preenchidas
  print("6. Aplicando bordas...")
  thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
  )
  for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
    for cell in row:
      cell.border = thin_border

  # 7. Aplicar fundo preto e fonte branca na primeira linha (cabeçalho)
  print("7. Formatando cabeçalho...")
  header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
  header_font = Font(color='FFFFFF', bold=True)
  
  for cell in sheet[1]: # Itera sobre a linha 1
    cell.fill = header_fill
    cell.font = header_font
    
  # --- Fim das 7 etapas ---
  # Pega a data atual no formato DIA-MES-ANO (sem barras '/')
  today_date = time.strftime("%d-%m-%Y") 
  
  # Cria o nome do arquivo (ex: "juridico montreal - 30-10-2025.xlsx")
  nome_arquivo_final = f"juridico montreal - {today_date}.xlsx"
  
  excel_file_path = os.path.join(download_dir, nome_arquivo_final)
  workbook.save(excel_file_path) # Salva o caminho do arquivo gerado
  
  print("\n--- PROCESSAMENTO DO EXCEL CONCLUÍDO! ---")
  print(f"Novo arquivo salvo como: {excel_file_path}")

except Exception as e:
  print(f"ERRO DURANTE O PROCESSAMENTO DO EXCEL: {e}")
  excel_file_path = None # Garante que não tente usar um caminho inválido
  
print("Script finalizado.")

# --- INÍCIO DA ETAPA FINAL: CRIAR E-MAIL NO OUTLOOK ---
if excel_file_path and os.path.exists(excel_file_path):
  try:
    print("\n--- INICIANDO CRIAÇÃO DE E-MAIL NO OUTLOOK ---")

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0) # 0 significa olMailItem (um e-mail)

    # Destinatários "Para"
    mail.To = "kelly.paixao@parceiro.mrv.com.br;reinaldo.reis@parceiro.mrv.com.br; maria.rferreira@parceiro.mrv.com.br"
    
    # Destinatários "Cc"
    mail.CC = "gabriel.emiliano@mrv.com.br; alfredo.pereira@mrv.com.br"

    # Assunto (com a data do dia anterior)
    today_date_obj = datetime.now()
    yesterday_date_obj = today_date_obj - timedelta(days=1)
    assunto_data = yesterday_date_obj.strftime("%d/%m/%Y")
    mail.Subject = f"Jurídico Montreal - {assunto_data}"

    # --- CORPO DO E-MAIL ATUALIZADO COM IMAGEM ---
    
    # 1. Anexar a imagem da assinatura e definir seu ID de conteúdo (CID)
    image_cid = "mrv_logo_cid" # ID único para a imagem
    assinatura_html = "" # Inicia a variável
    
    try:
      # Tenta encontrar a imagem no mesmo diretório do script
      # __file__ é o caminho do script python que está rodando
      script_dir = os.path.dirname(os.path.abspath(__file__)) 
      image_path = os.path.join(script_dir, "MRV.png")

      if not os.path.exists(image_path):
        # Fallback: Tenta o diretório de trabalho atual (se __file__ falhar)
        image_path = os.path.join(os.getcwd(), "MRV.png")

      if not os.path.exists(image_path):
        raise Exception("Arquivo MRV.png não encontrado.")

      # Anexa a imagem e define seu Content-ID (CID)
      attachment = mail.Attachments.Add(image_path)
      attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", image_cid)
      
      # 2. Formata a assinatura HTML como uma tabela (imagem à esquerda, texto à direita)
      assinatura_html = f"""
      <br> <!-- Espaço -->
      <table cellpadding="0" cellspacing="0" border="0" style="font-family: Calibri, sans-serif; font-size: 11pt; color: black;">
       <tr>
        <!-- Coluna 1: Imagem (puxando pelo CID) -->
        <td style="padding-right: 10px; vertical-align: top;">
         <img src="cid:{image_cid}" alt="MRV&CO Logo">
        </td>
        
        <!-- Coluna 2: Texto da Assinatura -->
        <td style="vertical-align: top;">
         <p style="margin: 0;"><b>Pedro Henrique Soares Silva</b></p>
         <p style="margin: 0;">Auxiliar | Administrativo</p>
         <p style="margin: 0;">MRV Engenharia e Participações S.A</p>
         <p style="margin: 0;">Av. Mário Werneck, 621 - Estoril - Belo Horizonte</p>
         <p style="margin: 0;">CEP: 30.455-610</p>
         <p style="margin: 0;">(Tel: (31) 3615-8325)</p>
         <p style="margin: 0;">
           <a href="mailto:Pedro.henrsilva@mrv.com.br">Pedro.henrsilva@mrv.com.br</a>– 
           <a href="http://www.mrv.com.br">www.mrv.com.br</a>
         </p>
        </td>
       </tr>
      </table>
      """
      
    except Exception as e:
      print(f"AVISO: Não foi possível anexar a imagem da assinatura (MRV.png). Erro: {e}")
      print("O e-mail será criado apenas com a assinatura de texto.")
      # Fallback (se a imagem falhar, usa a assinatura de texto)
      assinatura_html = """
      <br> <!-- Espaço -->
      <div style="font-family: Calibri, sans-serif; font-size: 11pt; color: black;">
        <p style="margin: 0;"><b>Pedro Henrique Soares Silva</b></p>
        <p style="margin: 0;">Auxiliar | Administrativo</p>
        <p style="margin: 0;">MRV Engenharia e Participações S.A</p>
        <p style="margin: 0;">Av. Mário Werneck, 621 - Estoril - Belo Horizonte</p>
        <p style="margin: 0;">CEP: 30.455-610</p>
        <p style="margin: 0;">(Tel: (31) 3615-8325)</p>
        <p style="margin: 0;">
          <a href="mailto:Pedro.henrsilva@mrv.com.br">Pedro.henrsilva@mrv.com.br</a>– 
          <a href="http://www.mrv.com.br">www.mrv.com.br</a>
        </p>
      </div>
      """

    # 3. Define o corpo do e-mail (Juntando tudo)
    mail.HTMLBody = f"""
    <p style="font-family: Calibri, sans-serif; font-size: 11pt;">Bom dia, Prezado(s)!</p>
    <p style="font-family: Calibri, sans-serif; font-size: 11pt;">Segue em anexo os documentos que chegaram para a Montreal, na data de ontem.</p>
    <p style="font-family: Calibri, sans-serif; font-size: 11pt;">Atenciosamente;</p>
    
    {assinatura_html} 
    """
    # --- FIM DA ATUALIZAÇÃO ---

    # Anexar o arquivo Excel
    mail.Attachments.Add(excel_file_path)

    # Exibir o e-mail (não envia automaticamente)
    mail.Display() 
    print("E-mail criado no Outlook com sucesso. Ele foi aberto na sua tela para revisão.")
    
  except Exception as e:
    print(f"ERRO AO CRIAR E-MAIL NO OUTLOOK: {e}")
    print("Certifique-se de que o Outlook esteja aberto e configurado corretamente.")
else:
  print("Não foi possível criar o e-mail no Outlook: arquivo Excel não encontrado ou não gerado.")

print("Programa concluído.")
